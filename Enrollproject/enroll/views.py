from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponseRedirect
from .forms import StudentForm,StudentForm2,PreEnrollForm,FormalEnrollForm,LoginForm
from .models import Student,Student2,YuBaoMing
import re
import openpyxl
from django.http import FileResponse
from openpyxl.styles import Font
from operator import itemgetter
import logging
import time
import pytz
def Login(request):
	if request.method == 'POST':
		form = LoginForm(request.POST)
		if form.is_valid():
			u = request.POST.get('name')
			p = request.POST.get('phone')
			user = YuBaoMing.objects.filter(name=u,phone=p)
			if user:
				user = YuBaoMing.objects.get(phone=p,name=u)
				form = FormalEnrollForm()
				major = user.major
				enroll_pay = user.is_pay
				money = user.money
				if enroll_pay==False:
					logging.debug(enroll_pay)
					return render(request,'enroll/weixin.html',{'message':'请完成预报名支付','money':100})
				request.session['name'] = u
				request.session['phone'] = p
				request.session['major'] =major
				data = [['姓名',user.name],['性别',user.sex],['手机',user.phone],['支付金额',user.money],['专业',user.major],['出发地',user.area],['订住酒店',user.need_dorm],['住宿天数',user.day_dorm],['大巴车',user.need_bus],['午餐',user.need_lunch]]
				if user.all_pay ==True:
					res =  render(request,'enroll/FormalEnroll.html',{'form':form,'name':u,'phone':p,'major':major,'end':'完成','message':'您已完成支付','st':data})
				elif user.is_enroll ==True:
					res =  render(request,'enroll/FormalEnroll.html',{'form':form,'name':u,'phone':p,'major':major,'message':'请完成支付','unend':'unend','st':data,'money':money})
				else:
					res =  render(request,'enroll/FormalEnroll.html',{'form':form,'name':u,'phone':p,'major':major})
				return res
			else:
				form = LoginForm()
				return render(request,'enroll/Login.html',{'message':'无账户信息,请完成预报名','form':form})
	form = LoginForm()
	return render(request, 'enroll/Login.html', {'form':form})
def FormalEnroll(request):
	try:
		name = request.session['name']
		phone = request.session['phone']
		major = request.session['major']
	except:
		form = LoginForm()
		return render(request, 'enroll/Login.html', {'form':form})
	if request.method == 'POST':
		form = FormalEnrollForm(request.POST)
		if form.is_valid():
			s = request.POST.get('sex')
			a = request.POST.get('area')
			need_bus = request.POST.get('need_bus')
			need_lunch = request.POST.get('need_lunch')
			need_dorm = request.POST.get('need_dorm')
			day_dorm = request.POST.get('day_dorm')
			money = request.POST.get('money')
			temp = YuBaoMing.objects.filter(name=name,phone=phone)
			data = [['姓名',name],['性别',s],['手机',phone],['支付金额',money],['专业',major],['出发地',a],['订住酒店',need_dorm],['住宿天数',day_dorm],['大巴车',need_bus],['午餐',need_lunch]]
			if temp:
				user = YuBaoMing.objects.get(name=name,phone=phone)
				if (user.all_pay == True):
					return render(request,'enroll/FormalEnroll.html',{'form':form,'message':'您已完成支付','end':'完成','st':data,'major':major})
				if (user.is_enroll == True):
					return render(request,'enroll/FormalEnroll.html',{'form':form,'message':'请完成支付','unend':'unend','st':data,'major':major})
				user.sex= s
				user.is_enroll = True
				user.area=a 
				user.need_bus= need_bus
				user.need_lunch= need_lunch
				user.need_dorm= need_dorm
				user.day_dorm = day_dorm;
				user.money = money;
				user.save()
				return render(request,'enroll/FormalEnroll.html',{'form':form,'success':'完成','message':'您的信息如下','st':data,'major':major,'money':money})
		else:
			return render(request, 'enroll/FormalEnroll.html', {'form':form,'name':name,'phone':phone,'major':major})
	else:
		form = FormalEnrollForm()
		return render(request, 'enroll/FormalEnroll.html', {'form':form,'name':name,'phone':phone,'major':major})
def Download4(request):
	wb = openpyxl.Workbook()
	sheet0 = wb.active
	sheet0.title = '管联'
	sheet1 = wb.create_sheet(title="医学")
	sheet2 = wb.create_sheet(title="建工")
	sheet3 = wb.create_sheet(title="其它")

	raw_data = [[],[],[],[]]
	value = [
			[['姓名','性别','手机','考点','专业','出发地','订住酒店','住宿天数','大巴车','午餐','预报名支付','金额','正式报名支付','信息提交时间']],
			[['姓名','性别','手机','考点','专业','出发地','订住酒店','住宿天数','大巴车','午餐','预报名支付','金额','正式报名支付','信息提交时间']],
			[['姓名','性别','手机','考点','专业','出发地','订住酒店','住宿天数','大巴车','午餐','预报名支付','金额','正式报名支付','信息提交时间']],
			[['姓名','性别','手机','考点','专业','出发地','订住酒店','住宿天数','大巴车','午餐','预报名支付','金额','正式报名支付','信息提交时间']],
				]
	a = YuBaoMing.objects.all().filter(area = '曹妃甸')
	shanghai = pytz.timezone("Asia/Shanghai")	
	for item in a:
		temp_time = shanghai.normalize(item.modifed_date.astimezone(shanghai))
		temp_time = str(temp_time).split('.')[0]
		if item.all_pay==True:
			all_pay = '是'
		else:
			all_pay = '否'
		if item.is_pay==True:
			is_pay = '是'
		else:
			is_pay = '否'
		temp = [item.name,item.sex,item.phone,item.exam_area,item.major,item.area,item.need_dorm,item.day_dorm,item.need_bus,item.need_lunch,is_pay,item.money,all_pay,temp_time]
		if item.major == "管联":
			raw_data[0].append(temp)
		elif item.major == "医学":
			raw_data[1].append(temp)
		elif item.major == "建工":
			raw_data[2].append(temp)
		else:
			raw_data[3].append(temp)
	for k in range(0,4):
		for temp in raw_data[k]:
			value[k].append(temp)

	sheetnames = wb.get_sheet_names()
	for k in range(0,4):
		ws = wb.get_sheet_by_name(sheetnames[k])
		for i in range(len(value[k])):
			for j in range(len(value[k][i])):
				ws.cell(row=i+1, column=j+1, value=str(value[k][i][j]))
	for k in range(0,4):
		ws = wb.get_sheet_by_name(sheetnames[k])
		for i in ['A','B','C','D','E','F','G','H','I','J','K','L','M']:
			ws.column_dimensions[i].width =25
	for  k in range(0,4):
		ws = wb.get_sheet_by_name(sheetnames[k])
		for column in ws:
			for cell in column:
				cell.font = Font(size=18)
	wb.save('CaoFeiDian.xlsx')
	file=open('CaoFeiDian.xlsx','rb')
	response =FileResponse(file)
	response['Content-Type']='application/octet-stream'
	response['Content-Disposition']='attachment;filename="CaoFeiDian.xlsx"'
	return response
def Download3(request):
	wb = openpyxl.Workbook()
	sheet0 = wb.active
	sheet0.title = '管联'
	sheet1 = wb.create_sheet(title="医学")
	sheet2 = wb.create_sheet(title="建工")
	sheet3 = wb.create_sheet(title="其它")

	raw_data = [[],[],[],[]]
	value = [
			[['姓名','性别','手机','考点','专业','出发地','订住酒店','住宿天数','大巴车','午餐','预报名支付','金额','正式报名支付','信息提交时间']],
			[['姓名','性别','手机','考点','专业','出发地','订住酒店','住宿天数','大巴车','午餐','预报名支付','金额','正式报名支付','信息提交时间']],
			[['姓名','性别','手机','考点','专业','出发地','订住酒店','住宿天数','大巴车','午餐','预报名支付','金额','正式报名支付','信息提交时间']],
			[['姓名','性别','手机','考点','专业','出发地','订住酒店','住宿天数','大巴车','午餐','预报名支付','金额','正式报名支付','信息提交时间']],
				]
	a = YuBaoMing.objects.all().filter(area = '市区')
	shanghai = pytz.timezone("Asia/Shanghai")	
	for item in a:
		temp_time = shanghai.normalize(item.modifed_date.astimezone(shanghai))
		temp_time = str(temp_time).split('.')[0]
		if item.all_pay==True:
			all_pay = '是'
		else:
			all_pay = '否'
		if item.is_pay==True:
			is_pay = '是'
		else:
			is_pay = '否'
		temp = [item.name,item.sex,item.phone,item.exam_area,item.major,item.area,item.need_dorm,item.day_dorm,item.need_bus,item.need_lunch,is_pay,item.money,all_pay,temp_time]
		if item.major == "管联":
			raw_data[0].append(temp)
		elif item.major == "医学":
			raw_data[1].append(temp)
		elif item.major == "建工":
			raw_data[2].append(temp)
		else:
			raw_data[3].append(temp)
	for k in range(0,4):
		for temp in raw_data[k]:
			value[k].append(temp)

	sheetnames = wb.get_sheet_names()
	for k in range(0,4):
		ws = wb.get_sheet_by_name(sheetnames[k])
		for i in range(len(value[k])):
			for j in range(len(value[k][i])):
				ws.cell(row=i+1, column=j+1, value=str(value[k][i][j]))
	for k in range(0,4):
		ws = wb.get_sheet_by_name(sheetnames[k])
		for i in ['A','B','C','D','E','F','G','H','I','J','K','L','M']:
			ws.column_dimensions[i].width =25
	for  k in range(0,4):
		ws = wb.get_sheet_by_name(sheetnames[k])
		for column in ws:
			for cell in column:
				cell.font = Font(size=18)
	wb.save('ShiQu.xlsx')
	file=open('ShiQu.xlsx','rb')
	response =FileResponse(file)
	response['Content-Type']='application/octet-stream'
	response['Content-Disposition']='attachment;filename="ShiQu.xlsx"'
	return response

def Download2(request):
	wb = openpyxl.Workbook()
	sheet = wb.active
	sheet.title = '文都考研现场确认报名表'
	value = [['姓名','性别','手机','QQ','乘车日期','乘车班次','单双程','学员','票价','12月份住宿等服务','专业','报考学院','报考专业','提交时间']]
	raw_data = []
	a = Student2.objects.all()
	shanghai = pytz.timezone("Asia/Shanghai")	
	for item in a:
		if item.is_need == True:
			is_need = '是'
		else:
			is_need = '否'
		temp_time = shanghai.normalize(item.modifed_date.astimezone(shanghai))
		temp_time = str(temp_time).split('.')[0]
		temp = [item.name,item.sex,item.phone,item.qq,item.ride_date,item.ride_time,item.is_return,item.is_enroll,item.price,is_need,item.major,item.obj_school,item.obj_major,temp_time]
		raw_data.append(temp)
	for item in raw_data:
		value.append(item)
	for i in range(len(value)):
		for j in range(len(value[i])):
			sheet.cell(row=i+1, column=j+1, value=str(value[i][j]))
	for i in ['A','B','C','D','E','F','G','H','I','J','K','L','M','N']:
		sheet.column_dimensions[i].width =25

	for column in sheet.columns:
		for cell in column:
			cell.font = Font(size=18)
	wb.save('xianchang.xlsx')
	file=open('xianchang.xlsx','rb')
	response =FileResponse(file)
	response['Content-Type']='application/octet-stream'
	response['Content-Disposition']='attachment;filename="xianchang.xlsx"'
	return response
def Download(request):
	wb = openpyxl.Workbook()
	sheet = wb.active
	sheet.title = '文都考研直通车报名表'
	value = [['姓名','手机','qq','是否学员','学院','专业','目标学校']]
	a = Student.objects.all()
	for item in a:
		if item.is_enroll == True:
			is_enroll = '是'
		else:
			is_enroll = '否'
		temp = [item.name,item.phone,item.qq,is_enroll,item.institute,item.major,item.obj_school]
		value.append(temp)
	for i in range(len(value)):
		for j in range(len(value[i])):
			sheet.cell(row=i+1, column=j+1, value=str(value[i][j]))
	for i in ['A','B','C','D','E','F','G']:
		sheet.column_dimensions[i].width = 30
	for column in sheet.columns:
		for cell in column:
			cell.font = Font(size=20)
	wb.save('./zhitongche.xlsx')
	file=open('./zhitongche.xlsx','rb')
	response =FileResponse(file)
	response['Content-Type']='application/octet-stream'
	response['Content-Disposition']='attachment;filename="zhitongche.xlsx"'
	return response
def index(request):
	return render(request, 'enroll/index.html')
def Enroll(request):
	if request.method == 'POST':
		form = StudentForm(request.POST)
		if form.is_valid():
			context = {'form':form}
			u = request.POST.get('name')
			q = request.POST.get('qq',None)
			o = request.POST.get('obj_school',None)
			i = request.POST.get('institute',None)
			m = request.POST.get('major',None)
			p = request.POST.get('phone')
			is_en = request.POST.get('is_enroll',False)
			temp = Student.objects.filter(name=u,phone=p)
			if temp:
				context.update({'message':'已报名,请勿重复提交'})
				return render(request, 'enroll/enroll.html', context=context)

			reg = re.compile(r'^1[0-9]{10}$')	
			if reg.match(p) == None:
				context.update({'message':'手机号请正确填写'})
				return render(request, 'enroll/enroll.html', context=context)

			if is_en !=False:
				is_en = True
			Student.objects.create(
			name = u,
			qq = q,
			obj_school = o,
			institute = i,
			major = m,
			phone = p,
			is_enroll = is_en,
			)
			return render(request, 'enroll/enroll_ok.html', {'message':'[%s]同学报名成功,点击返回'%request.POST['name']})
	form = StudentForm()
	return render(request, 'enroll/enroll.html', {'form':form})
def PreEnroll(request,*args,**kwargs):
	if request.method == 'POST':
		form = PreEnrollForm(request.POST)
		if form.is_valid():
			context = {'form':form}
			u = request.POST.get('name')
			m = request.POST.get('major')
			p = request.POST.get('phone')
			a = request.POST.get('area')
			need_bus = request.POST.get('need_bus')
			need_dorm = request.POST.get('need_dorm')
			temp = YuBaoMing.objects.filter(name=u,phone=p)
			data = [['姓名',u],['手机',p],['专业',m],['出发地',a],['订住酒店',need_dorm],['大巴车',need_bus]]
			if temp:
				return render(request,'enroll/PreEnroll.html',{'st':data,'money':100,'message1':'已提交信息,请勿重复提交'})
			reg = re.compile(r'^1[0-9]{10}$')	
			if reg.match(p) == None:
				context.update({'message2':'手机号请正确填写'})
				return render(request, 'enroll/PreEnroll.html', context=context)
			YuBaoMing.objects.create(
			name = u,
			major = m,
			phone = p,
			is_pay= False,
			area = a,
			need_bus = need_bus,
			need_dorm = need_dorm,
			)
			return render(request,'enroll/PreEnroll.html',{'st':data,'money':100})
	form = PreEnrollForm()
	return render(request, 'enroll/PreEnroll.html', {'form':form})
def Enroll2(request,*args,**kwargs):
	logging.debug(args)
	if request.method == 'POST':
		form = StudentForm2(request.POST)
		if form.is_valid():
			context = {'form':form}
			u = request.POST.get('name')
			q = request.POST.get('qq')
			o = request.POST.get('obj_school')
			om = request.POST.get('obj_major')
			m = request.POST.get('major')
			p = request.POST.get('phone')
			rd = request.POST.get('ride_date')
			rt = request.POST.get('ride_time')
			is_re = request.POST.get('is_return')
			is_en = request.POST.get('is_enroll')
			sex = request.POST.get('sex')
			is_ne = request.POST.get('is_need',False)
			try:
				if is_re == '单程':
					pr = 25
				else:
					pr = 45
				if is_en == '是':
					pr -=5
			except:
				pass
			if is_re == '0':
				context.update({'message':'请填写乘车等相关信息'})
				return render(request, 'enroll/enroll2.html', context=context)
			temp = Student2.objects.filter(name=u,phone=p)
			if temp:
				rd = temp[0].ride_date
				rt =temp[0].ride_time
				ph = temp[0].phone
				name = temp[0].name
				context.update({'message':'已提交信息,请勿重复提交','ph':ph,'rd':rd,'rt':rt,'name':name})
				return render(request, 'enroll/enroll2.html', context=context)
			reg = re.compile(r'^1[0-9]{10}$')	
			if reg.match(p) == None:
				context.update({'message':'手机号请正确填写'})
				return render(request, 'enroll/enroll2.html', context=context)
			if is_ne !=False:
				is_ne = True
				ne  = '需要'
			else:	
				ne = '不需要'

			data = [['姓名',u],['性别',sex],['手机',p],['QQ',q],['乘车日期',rd],['乘车班次',rt],['单双程',is_re],['是否学员',is_en],['票价',pr],['十二月份是否需要住宿等服务',ne],['专业',m],['目标学校',o],['目标专业',om]]

			logging.debug(pr)
			Student2.objects.create(
			name = u,
			qq = q,
			obj_school = o,
			obj_major = om,
			ride_date = rd,
			ride_time = rt,
			major = m,
			phone = p,
			is_return = is_re,
			is_need = is_ne,
			sex= sex,
			is_enroll=is_en,
			price = pr,
			)
			rd = rd
			rt = rt
			ph = p
			name = u
			return render(request, 'enroll/enroll2_ok.html', {'message':'您的信息如下','st':data,'money':pr})
	form = StudentForm2()
	return render(request, 'enroll/enroll2.html', {'form':form})
def zhifubao(request,money):
	return render(request,'enroll/zhifubao_%s.html'%(str(money)))
def weixin(request,money):
	return render(request,'enroll/weixin.html',{'money':money})
